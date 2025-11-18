"""Service for exporting reports to Excel and PDF formats."""

from typing import List, Dict, Any, Optional
from datetime import datetime
from io import BytesIO
from decimal import Decimal


def format_currency(amount: Optional[float]) -> str:
    """Format amount as LKR currency."""
    if amount is None:
        return ""
    return f"LKR {amount:,.2f}"


def to_excel_value(value: Any) -> Any:
    """Convert value to Excel-compatible type."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


class ExcelExporter:
    """
    Excel export service using openpyxl.

    Note: Requires openpyxl package. Install with: pip install openpyxl
    """

    @staticmethod
    def export_acquisition_ledger(
        bikes: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        Export acquisition ledger to Excel (November notebook format).

        Args:
            bikes: List of bike dictionaries
            filters: Optional filter parameters used for the report

        Returns:
            BytesIO object containing Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Acquisition Ledger"

        # Title
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = "Bike Acquisition Ledger"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Filters info
        if filters:
            ws.merge_cells('A2:K2')
            filter_cell = ws['A2']
            filter_parts = []
            if filters.get('company_id'):
                filter_parts.append(f"Company: {filters['company_id']}")
            if filters.get('branch_id'):
                filter_parts.append(f"Branch: {filters['branch_id']}")
            if filters.get('start_date') or filters.get('end_date'):
                filter_parts.append(
                    f"Date: {filters.get('start_date', '')} to {filters.get('end_date', '')}"
                )
            filter_cell.value = f"Filters: {' | '.join(filter_parts)}" if filter_parts else "All Records"
            filter_cell.font = Font(italic=True)

        # Headers
        headers = [
            "Stock Number",
            "Date",
            "Company",
            "Branch",
            "Brand",
            "Model",
            "Year",
            "Purchase Price",
            "Supplier",
            "Procured By",
            "Notes"
        ]

        header_row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, bike in enumerate(bikes, start=header_row + 1):
            ws.cell(row=row_idx, column=1, value=bike.get('current_stock_number', ''))
            ws.cell(row=row_idx, column=2, value=bike.get('procurement_date', ''))
            ws.cell(row=row_idx, column=3, value=bike.get('company_id', ''))
            ws.cell(row=row_idx, column=4, value=bike.get('current_branch_id', ''))
            ws.cell(row=row_idx, column=5, value=bike.get('brand', ''))
            ws.cell(row=row_idx, column=6, value=bike.get('model', ''))
            ws.cell(row=row_idx, column=7, value=bike.get('year', ''))

            # Format currency
            price_cell = ws.cell(row=row_idx, column=8)
            price_cell.value = to_excel_value(bike.get('base_purchase_price'))
            price_cell.number_format = '#,##0.00'

            ws.cell(row=row_idx, column=9, value=bike.get('supplier_name', ''))
            ws.cell(row=row_idx, column=10, value=bike.get('procured_by', ''))
            ws.cell(row=row_idx, column=11, value=bike.get('procurement_notes', ''))

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_cost_summary(
        bikes: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        Export cost summary to Excel (summery.xlsx format).

        Args:
            bikes: List of bike dictionaries with cost details
            filters: Optional filter parameters

        Returns:
            BytesIO object containing Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")

        wb = Workbook()
        ws = wb.active
        ws.title = "Cost Summary"

        # Title
        ws.merge_cells('A1:M1')
        title_cell = ws['A1']
        title_cell.value = "Bike Cost Summary Report"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Headers
        headers = [
            "Stock Number",
            "Brand/Model",
            "Year",
            "Company",
            "Branch",
            "Status",
            "Purchase Price",
            "Repair Cost",
            "Branch Expenses",
            "Total Cost",
            "Selling Price",
            "Profit/Loss",
            "P/L %"
        ]

        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        total_purchase = 0
        total_repair = 0
        total_expenses = 0
        total_cost = 0
        total_selling = 0
        total_profit = 0

        for row_idx, bike in enumerate(bikes, start=header_row + 1):
            ws.cell(row=row_idx, column=1, value=bike.get('current_stock_number', ''))
            ws.cell(row=row_idx, column=2, value=f"{bike.get('brand', '')} {bike.get('model', '')}")
            ws.cell(row=row_idx, column=3, value=bike.get('year', ''))
            ws.cell(row=row_idx, column=4, value=bike.get('company_id', ''))
            ws.cell(row=row_idx, column=5, value=bike.get('current_branch_id', ''))
            ws.cell(row=row_idx, column=6, value=bike.get('status', ''))

            # Financial columns with formatting
            purchase = to_excel_value(bike.get('base_purchase_price', 0))
            repair = to_excel_value(bike.get('total_repair_cost', 0))
            expenses = to_excel_value(bike.get('total_branch_expenses', 0))
            total = purchase + repair + expenses
            selling = to_excel_value(bike.get('selling_price'))
            profit = selling - total if selling else None

            ws.cell(row=row_idx, column=7, value=purchase).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=8, value=repair).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=9, value=expenses).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=10, value=total).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=11, value=selling or '').number_format = '#,##0.00'

            # Profit/Loss with conditional formatting
            if profit is not None:
                profit_cell = ws.cell(row=row_idx, column=12, value=profit)
                profit_cell.number_format = '#,##0.00'
                if profit < 0:
                    profit_cell.font = Font(color="FF0000")  # Red for loss
                elif profit > 0:
                    profit_cell.font = Font(color="00FF00")  # Green for profit

                # P/L percentage
                pl_pct = (profit / total * 100) if total > 0 else 0
                pct_cell = ws.cell(row=row_idx, column=13, value=pl_pct)
                pct_cell.number_format = '0.00%'

            # Accumulate totals
            total_purchase += purchase
            total_repair += repair
            total_expenses += expenses
            total_cost += total
            if selling:
                total_selling += selling
            if profit:
                total_profit += profit

        # Totals row
        totals_row = len(bikes) + header_row + 1
        ws.cell(row=totals_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=totals_row, column=7, value=total_purchase).number_format = '#,##0.00'
        ws.cell(row=totals_row, column=8, value=total_repair).number_format = '#,##0.00'
        ws.cell(row=totals_row, column=9, value=total_expenses).number_format = '#,##0.00'
        ws.cell(row=totals_row, column=10, value=total_cost).number_format = '#,##0.00'
        ws.cell(row=totals_row, column=11, value=total_selling).number_format = '#,##0.00'
        ws.cell(row=totals_row, column=12, value=total_profit).number_format = '#,##0.00'

        # Make totals row bold
        for col in range(1, 14):
            ws.cell(row=totals_row, column=col).font = Font(bold=True)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output


class PDFExporter:
    """
    PDF export service using reportlab.

    Note: Requires reportlab package. Install with: pip install reportlab
    """

    @staticmethod
    def export_commission_report(
        commissions: List[Dict[str, Any]],
        filters: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        Export commission report to PDF.

        Args:
            commissions: List of commission dictionaries
            filters: Optional filter parameters

        Returns:
            BytesIO object containing PDF file
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        except ImportError:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")

        output = BytesIO()

        # Create PDF
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph("Commission Report", title_style))

        # Filters
        if filters:
            filter_text = []
            if filters.get('branch_id'):
                filter_text.append(f"Branch: {filters['branch_id']}")
            if filters.get('start_date') or filters.get('end_date'):
                filter_text.append(
                    f"Period: {filters.get('start_date', '')} to {filters.get('end_date', '')}"
                )

            if filter_text:
                elements.append(Paragraph(
                    ' | '.join(filter_text),
                    styles['Normal']
                ))
                elements.append(Spacer(1, 12))

        # Table data
        data = [[
            "Date",
            "Stock #",
            "Bike",
            "Branch",
            "Sale Price",
            "Profit",
            "Commission",
            "Type"
        ]]

        total_sale_price = 0
        total_profit = 0
        total_commission = 0

        for comm in commissions:
            data.append([
                comm.get('sale_date', '')[:10] if comm.get('sale_date') else '',
                comm.get('stock_number', ''),
                f"{comm.get('bike_brand', '')} {comm.get('bike_model', '')}",
                comm.get('branch_id', ''),
                format_currency(comm.get('sale_price')),
                format_currency(comm.get('profit')),
                format_currency(comm.get('commission_amount')),
                comm.get('commission_type', '')
            ])

            total_sale_price += comm.get('sale_price', 0)
            total_profit += comm.get('profit', 0)
            total_commission += comm.get('commission_amount', 0)

        # Totals row
        data.append([
            "TOTAL",
            "",
            "",
            "",
            format_currency(total_sale_price),
            format_currency(total_profit),
            format_currency(total_commission),
            ""
        ])

        # Create table
        table = Table(data, colWidths=[
            0.8*inch,  # Date
            0.9*inch,  # Stock #
            1.5*inch,  # Bike
            0.7*inch,  # Branch
            1.0*inch,  # Sale Price
            0.9*inch,  # Profit
            1.0*inch,  # Commission
            0.7*inch   # Type
        ])

        # Table style
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')]),

            # Totals row
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d0d0d0')),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(table)

        # Summary section
        elements.append(Spacer(1, 20))
        summary_style = styles['Normal']
        elements.append(Paragraph(
            f"<b>Report Summary:</b> {len(commissions)} commission entries",
            summary_style
        ))
        elements.append(Paragraph(
            f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            summary_style
        ))

        # Build PDF
        doc.build(elements)
        output.seek(0)

        return output
